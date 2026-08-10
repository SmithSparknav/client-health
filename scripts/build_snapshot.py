#!/usr/bin/env python3
"""Build the published ClientPulse snapshot from approved weekly source files."""

import argparse
import json
import re
from datetime import date, datetime, timedelta
from html.parser import HTMLParser
from pathlib import Path

from openpyxl import load_workbook

DAY_SECONDS = 86_400


def clean(value):
    return "" if value is None else str(value).strip()


def normalize(value):
    return re.sub(r"[^a-z0-9]+", " ", clean(value).lower().replace("&", "and")).strip()


ALIASES = {
    normalize("SeedSpark Client"): normalize("SeedSpark"),
    normalize("SparkNav (internal)"): normalize("SparkNav"),
    normalize("JenCon Builders"): normalize("Jencon Builders"),
    normalize("St. Amand & Efird"): normalize("St. Amand and Efird"),
}

APPROVED_TIER_OVERRIDES = {
    "AccruePartners": "Tier 1",
    "Blue Dot Readi-Mix": "Tier 1",
    "Carolina Ingredients": "Tier 1",
    "Enviro-Master Services": "Tier 1",
    "Mechanical Systems & Services (MSS)": "Tier 1",
    "Red Moon Marketing": "Tier 1",
}


class TableParser(HTMLParser):
    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.rows, self.row, self.cell, self.in_cell = [], [], [], False

    def handle_starttag(self, tag, attrs):
        if tag == "tr":
            self.row = []
        elif tag in {"td", "th"}:
            self.cell, self.in_cell = [], True

    def handle_data(self, data):
        if self.in_cell:
            self.cell.append(data)

    def handle_endtag(self, tag):
        if tag in {"td", "th"} and self.in_cell:
            self.row.append("".join(self.cell).strip())
            self.in_cell = False
        elif tag == "tr" and self.row:
            self.rows.append(self.row)


def workbook_rows(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    return [list(row) for row in wb.active.iter_rows(values_only=True)]


def html_rows(path):
    parser = TableParser()
    parser.feed(Path(path).read_text(errors="ignore"))
    return parser.rows


def as_datetime(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    text = clean(value)
    for fmt in ("%b %d, %Y", "%B %d, %Y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None


def report_date(rows, fallback=None):
    for row in rows[:8]:
        for value in row:
            if isinstance(value, (date, datetime)):
                return as_datetime(value)
            match = re.search(r"(?:as of\s+)?([A-Z][a-z]+\s+\d{1,2},\s+\d{4})", clean(value))
            if match:
                parsed = as_datetime(match.group(1))
                if parsed:
                    return parsed
    return fallback


def find_header(rows, required):
    wanted = [normalize(item) for item in required]
    for index, row in enumerate(rows):
        values = [normalize(item) for item in row]
        if all(item in values for item in wanted):
            return index
    raise ValueError(f"Required columns not found: {', '.join(required)}")


def parse_ar(path):
    rows = workbook_rows(path)
    header_index = find_header(rows, ["CURRENT", "1 - 30", "31 - 60", "61 - 90", "91 AND OVER"])
    header = [normalize(item) for item in rows[header_index]]
    columns = {label: header.index(normalize(label)) for label in ["CURRENT", "1 - 30", "31 - 60", "61 - 90", "91 AND OVER"]}
    accounts = {}
    for row in rows[header_index + 1:]:
        name = clean(row[0] if row else "")
        if normalize(name) == "total":
            break
        if not name:
            continue
        def number(label):
            value = row[columns[label]] if columns[label] < len(row) else 0
            return float(value or 0)
        balances = {"current": number("CURRENT"), "oneThirty": number("1 - 30"), "thirtySixty": number("31 - 60"), "sixtyNinety": number("61 - 90"), "ninetyPlus": number("91 AND OVER")}
        bucket = "90+" if balances["ninetyPlus"] > 0 else "61-90" if balances["sixtyNinety"] > 0 else "31-60" if balances["thirtySixty"] > 0 else "Current"
        amount = {"90+": balances["ninetyPlus"], "61-90": balances["sixtyNinety"], "31-60": balances["thirtySixty"], "Current": 0}[bucket]
        accounts[name] = {"bucket": bucket, "amount": amount, "balances": balances}
    as_of = report_date(rows)
    if not as_of:
        raise ValueError("AR report date could not be determined")
    return {"fileName": Path(path).name, "asOf": as_of.date().isoformat(), "accounts": accounts}


def parse_open(path):
    rows = workbook_rows(path)
    header_index = find_header(rows, ["Date Received", "Ticket Number", "Client Name", "Priority"])
    header = [normalize(item) for item in rows[header_index]]
    col = lambda label: header.index(normalize(label))
    tickets = []
    for row in rows[header_index + 1:]:
        number, client = clean(row[col("Ticket Number")]), clean(row[col("Client Name")])
        received = as_datetime(row[col("Date Received")])
        if number and client and received:
            tickets.append({"number": number, "client": client, "received": received, "priority": clean(row[col("Priority")])})
    as_of = report_date(rows, max(item["received"] for item in tickets))
    oldest = min(item["received"] for item in tickets)
    span = max(0, int((as_of - oldest).total_seconds() // DAY_SECONDS))
    return {"fileName": Path(path).name, "asOf": as_of.date().isoformat(), "tickets": tickets, "diagnostics": {"rowCount": len(tickets), "oldestReceived": oldest.date().isoformat(), "spanDays": span, "possibleTruncation": len(tickets) >= 1100 and span <= 35}}


def parse_volume(path):
    rows = html_rows(path)
    tickets, active = [], None
    for row in rows:
        normalized = [normalize(item) for item in row]
        if all(item in normalized for item in ["task or ticket number", "client name", "status", "priority"]):
            # Autotask repeats the company grouping label as the first header
            # cell while the ticket rows contain only the seven report columns.
            active = normalized[1:] if normalized and normalized[0].startswith("client name ") else normalized
            continue
        if not active:
            continue
        def value(label):
            index = active.index(label)
            return clean(row[index] if index < len(row) else "")
        number, client = value("task or ticket number"), value("client name")
        if re.match(r"^T\d+", number, re.I) and client:
            tickets.append({"number": number, "client": client, "status": value("status"), "priority": value("priority")})
    text = " ".join(clean(item) for row in rows[:3] for item in row)
    dates = [as_datetime(item).date().isoformat() for item in re.findall(r"\d{2}/\d{2}/\d{4}", text)]
    return {"fileName": Path(path).name, "periodStart": dates[0] if dates else None, "periodEnd": dates[1] if len(dates) > 1 else dates[0] if dates else None, "tickets": tickets}


def build(master, ar, opened, volume):
    canonical = {normalize(item["name"]): item["name"] for item in master}
    def resolve(name):
        key = normalize(name)
        return canonical.get(ALIASES.get(key, key))
    def group(rows, key):
        matched, unmatched = {}, set()
        for row in rows:
            name = resolve(row[key])
            if name:
                matched.setdefault(name, []).append(row)
            else:
                unmatched.add(row[key])
        return matched, unmatched
    ar_rows = [{"sourceName": name, **value} for name, value in ar["accounts"].items()]
    ar_grouped, ar_unmatched = group(ar_rows, "sourceName")
    open_grouped, open_unmatched = group(opened["tickets"], "client")
    volume_grouped, volume_unmatched = group(volume["tickets"], "client")
    # Browser calculation parses the ISO report date at UTC midnight while
    # Excel ticket dates are local. Match that established ClientPulse result.
    as_of = as_datetime(opened["asOf"]) - timedelta(hours=4)
    provisional = opened["diagnostics"]["possibleTruncation"]
    clients = {}
    for item in master:
        name = item["name"]
        ar_record = (ar_grouped.get(name) or [None])[0]
        open_rows, volume_rows = open_grouped.get(name, []), volume_grouped.get(name, [])
        age = max((max(0, int((as_of - ticket["received"]).total_seconds() // DAY_SECONDS)) for ticket in open_rows), default=None)
        bucket = ar_record["bucket"] if ar_record else "Current"
        payment = {"Current": 20, "31-60": 10, "61-90": 4, "90+": 0}[bucket]
        service = 40 if age is None else 4 if age >= 31 else 12 if age >= 21 else 24 if age >= 11 else 34 if age >= 5 else 40
        support = 30 if age is None else 5 if age >= 31 else 10 if age >= 21 else 18 if age >= 11 else 28
        score, caps = payment + service + support + 10, []
        for cap_id, applies, cap in [("ar_90_plus", bucket == "90+", 50), ("ar_61_90", bucket == "61-90", 62), ("ar_31_60", bucket == "31-60", 74), ("ticket_30_day", age is not None and age >= 30, 69), ("ticket_21_day", age is not None and age >= 21, 79)]:
            if applies and score > cap:
                score, caps = cap, caps + [cap_id]
        clients[name] = {"score": score, "band": "Healthy" if score >= 85 else "Watch" if score >= 70 else "At Risk", "scoreStatus": "Provisional" if provisional else "Verified", "dataState": "Partial Data" if provisional else "Complete", "pillars": {"paymentHealth": payment, "serviceResponsiveness": service, "supportDemand": support, "engagement": 10}, "ar": {"bucket": bucket, "amount": ar_record["amount"] if ar_record else 0, "asOf": ar["asOf"], "totalOutstanding": max(0, sum(ar_record["balances"].values())) if ar_record else 0}, "tickets": {"openCount": len(open_rows), "oldestOpenAgeDays": age, "volumeCount": len(volume_rows), "asOf": opened["asOf"]}, "capsApplied": caps}
    def rank10(name, field):
        value = field(clients[name])
        if value <= 0:
            return 1
        positives = sorted(field(client) for client in clients.values() if field(client) > 0)
        upper = sum(item <= value for item in positives)
        return min(10, 1 + -(-upper * 9 // len(positives)))
    tiers = []
    for name in clients:
        financial, ticket = rank10(name, lambda c: c["ar"]["totalOutstanding"]), rank10(name, lambda c: c["tickets"]["volumeCount"])
        tiers.append({"name": name, "financialRank": financial, "ticketVolumeRank": ticket, "tierScore": round(financial * 35 / 65 + ticket * 30 / 65, 2)})
    tiers.sort(key=lambda row: (-row["tierScore"], -row["financialRank"], -row["ticketVolumeRank"], row["name"].lower()))
    first, second = -(-len(tiers) // 3), -(-(len(tiers) * 2) // 3)
    for index, row in enumerate(tiers):
        row["tier"] = "Tier 1" if index < first else "Tier 2" if index < second else "Tier 3"
        row["method"] = "Available Data Tier Score v1"
        clients[row["name"]]["tiering"] = row
    for name, tier in APPROVED_TIER_OVERRIDES.items():
        if name not in clients:
            raise ValueError(f"Approved tier override client is missing: {name}")
        tiering = clients[name]["tiering"]
        tiering.update({"calculatedTier": tiering["tier"], "tier": tier, "override": True, "tierSource": "Approved manual assignment", "method": "Approved Tier Override"})
    return {"schemaVersion": "2.0", "generatedAt": datetime.now().astimezone().isoformat(), "scoringModel": "ClientPulse v1", "tieringModel": {"name": "Available Data Tier Score v1", "financialFactor": "Total outstanding AR balance", "financialWeight": 35 / 65, "ticketVolumeWeight": 30 / 65, "populationSplit": "Top, middle, and bottom thirds"}, "sources": {"ar": {"fileName": ar["fileName"], "asOf": ar["asOf"], "matchedClients": len(ar_grouped)}, "openTickets": {"fileName": opened["fileName"], "asOf": opened["asOf"], "matchedClients": len(open_grouped), **opened["diagnostics"]}, "ticketVolume": {"fileName": volume["fileName"], "periodStart": volume["periodStart"], "periodEnd": volume["periodEnd"], "matchedClients": len(volume_grouped), "rowCount": len(volume["tickets"])}}, "clients": clients, "exceptions": {"unmatchedAr": sorted(ar_unmatched), "unmatchedOpenTickets": sorted(open_unmatched), "unmatchedTicketVolume": sorted(volume_unmatched)}, "warnings": ["Open Tickets report may be truncated: high row count with only a short received-date span. Scores are provisional."] if provisional else []}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--clients", required=True)
    parser.add_argument("--ar", required=True)
    parser.add_argument("--open-tickets", required=True)
    parser.add_argument("--ticket-volume", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()
    master = json.loads(Path(args.clients).read_text())["clients"]
    if len(master) != 175 or len({item["name"] for item in master}) != 175:
        raise SystemExit("Master client population must contain 175 unique names")
    snapshot = build(master, parse_ar(args.ar), parse_open(args.open_tickets), parse_volume(args.ticket_volume))
    Path(args.output).write_text(json.dumps(snapshot, indent=2) + "\n")
    print(f"Published snapshot built for {len(snapshot['clients'])} clients")


if __name__ == "__main__":
    main()
